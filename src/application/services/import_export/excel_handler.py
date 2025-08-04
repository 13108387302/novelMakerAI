#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel格式处理器

处理Excel格式的项目和文档导入导出
"""

from pathlib import Path
from typing import List, Optional
from datetime import datetime

from .base import BaseFormatHandler, ExportOptions, ImportOptions
from src.domain.entities.project import Project
from src.domain.entities.document import Document, DocumentType
from src.shared.utils.logger import get_logger

logger = get_logger(__name__)

# 检查Excel库是否可用
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl库不可用，Excel格式处理器将无法正常工作")


class ExcelFormatHandler(BaseFormatHandler):
    """
    Excel格式处理器

    处理Excel格式的项目和文档导入导出操作。
    使用openpyxl库生成结构化的Excel工作簿。

    实现方式：
    - 使用openpyxl库处理Excel文件
    - 创建多个工作表组织不同类型的数据
    - 支持样式和格式化
    - 提供数据验证和导入功能
    """

    def get_supported_extensions(self) -> List[str]:
        """
        获取支持的文件扩展名

        Returns:
            List[str]: 支持的文件扩展名列表
        """
        return ['.xlsx', '.xls']

    def get_format_name(self) -> str:
        """
        获取格式名称

        Returns:
            str: 格式的显示名称
        """
        return "Excel工作簿"

    async def _do_export_project(self, project: Project, output_path: Path, options: ExportOptions) -> bool:
        """导出项目为Excel格式"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl库不可用，无法导出Excel格式")
            return False
            
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建项目信息工作表
            self._create_project_info_sheet(wb, project, options)
            
            # 创建文档列表工作表
            if options.include_documents:
                await self._create_documents_sheet(wb, project, options)
            
            # 创建统计信息工作表
            if options.include_statistics:
                await self._create_statistics_sheet(wb, project, options)
            
            # 保存工作簿
            wb.save(str(output_path))
            
            logger.info(f"项目已导出为Excel: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            return False

    def _create_project_info_sheet(self, wb, project: Project, options: ExportOptions):
        """创建项目信息工作表"""
        ws = wb.create_sheet("项目信息")
        
        # 设置标题样式
        title_font = Font(size=16, bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置标题
        ws['A1'] = "项目信息"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = title_alignment
        ws.merge_cells('A1:B1')
        
        # 设置行高
        ws.row_dimensions[1].height = 30
        
        # 添加项目信息
        info_data = [
            ("项目名称", project.name),
            ("项目描述", getattr(project, 'description', '')),
            ("创建时间", getattr(project, 'created_at', '').strftime('%Y-%m-%d %H:%M:%S') if hasattr(project, 'created_at') and project.created_at else ''),
            ("导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("导出工具", "AI小说编辑器"),
        ]
        
        # 设置数据样式
        label_font = Font(bold=True)
        
        for i, (label, value) in enumerate(info_data, 2):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = label_font
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50

    async def _create_documents_sheet(self, wb, project: Project, options: ExportOptions):
        """创建文档列表工作表"""
        ws = wb.create_sheet("文档列表")
        
        # 设置标题
        headers = ["序号", "文档标题", "文档类型", "字符数", "词数", "创建时间", "内容预览"]
        
        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置行高
        ws.row_dimensions[1].height = 25
        
        try:
            # 获取文档
            documents = await self.service.document_repository.get_by_project_id(project.id)
            
            # 添加文档数据
            for i, doc in enumerate(documents, 2):
                char_count = len(doc.content)
                word_count = len(doc.content.split())
                preview = doc.content[:100] + "..." if len(doc.content) > 100 else doc.content
                created_time = doc.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(doc, 'created_at') and doc.created_at else ''
                
                row_data = [
                    i - 1,  # 序号
                    doc.title,
                    str(doc.type),
                    char_count,
                    word_count,
                    created_time,
                    preview
                ]
                
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=i, column=col, value=value)
                    
        except Exception as e:
            logger.warning(f"获取文档列表失败: {e}")
            ws.cell(row=2, column=1, value="无法获取文档列表")
        
        # 调整列宽
        column_widths = [8, 30, 15, 12, 12, 20, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    async def _create_statistics_sheet(self, wb, project: Project, options: ExportOptions):
        """创建统计信息工作表"""
        ws = wb.create_sheet("统计信息")
        
        # 设置标题
        ws['A1'] = "项目统计信息"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 30
        
        # 计算统计信息
        total_chars = 0
        total_words = 0
        doc_count = 0
        
        try:
            documents = await self.service.document_repository.get_by_project_id(project.id)
            doc_count = len(documents)
            
            for doc in documents:
                total_chars += len(doc.content)
                total_words += len(doc.content.split())
                
        except Exception as e:
            logger.warning(f"计算统计信息失败: {e}")
        
        # 添加统计数据
        stats_data = [
            ("文档数量", doc_count),
            ("总字符数", f"{total_chars:,}"),
            ("总词数", f"{total_words:,}"),
            ("平均每文档字符数", f"{total_chars // doc_count:,}" if doc_count > 0 else "0"),
            ("平均每文档词数", f"{total_words // doc_count:,}" if doc_count > 0 else "0"),
        ]
        
        label_font = Font(bold=True)
        
        for i, (label, value) in enumerate(stats_data, 2):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = label_font
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    async def _do_import_project(self, input_path: Path, options: ImportOptions) -> Optional[Project]:
        """从Excel导入项目"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl库不可用，无法导入Excel格式")
            return None
            
        try:
            # 验证文件扩展名
            if not self._validate_file_extension(input_path, self.get_supported_extensions()):
                logger.error(f"不支持的文件格式: {input_path.suffix}")
                return None

            # 读取Excel文件
            wb = openpyxl.load_workbook(str(input_path))
            
            # 尝试从项目信息工作表获取项目信息
            project_name = input_path.stem
            project_description = ""
            
            if "项目信息" in wb.sheetnames:
                ws = wb["项目信息"]
                for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                    if row[0] == "项目名称" and row[1]:
                        project_name = row[1]
                    elif row[0] == "项目描述" and row[1]:
                        project_description = row[1]
            
            # 创建项目
            project = Project(
                id="",  # 将由仓储分配
                name=project_name,
                description=project_description or f"从Excel文件导入: {input_path.name}"
            )
            
            # 尝试从文档列表工作表导入文档
            if "文档列表" in wb.sheetnames:
                await self._import_documents_from_excel(wb["文档列表"], project)
            
            logger.info(f"项目已从Excel导入: {project.name}")
            return project
            
        except Exception as e:
            logger.error(f"从Excel导入项目失败: {e}")
            return None

    async def _import_documents_from_excel(self, ws, project: Project):
        """从Excel工作表导入文档"""
        try:
            # 跳过标题行，从第二行开始读取
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[1]:  # 确保有标题
                    title = row[1]
                    doc_type = row[2] if len(row) > 2 and row[2] else "chapter"
                    preview = row[6] if len(row) > 6 and row[6] else ""
                    
                    # 创建文档
                    document = Document(
                        id="",  # 将由仓储分配
                        title=title,
                        content=preview,  # 使用预览内容作为文档内容
                        type=DocumentType(doc_type) if doc_type in [t.value for t in DocumentType] else DocumentType.CHAPTER
                    )
                    
                    # 保存文档
                    await self.service.document_repository.save(document)
                    
        except Exception as e:
            logger.error(f"从Excel导入文档失败: {e}")

    async def _do_export_document(self, document: Document, output_path: Path, options: ExportOptions) -> bool:
        """导出文档为Excel格式"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl库不可用，无法导出Excel格式")
            return False
            
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "文档内容"
            
            # 设置文档标题
            ws['A1'] = document.title
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells('A1:B1')
            ws.row_dimensions[1].height = 30
            
            # 添加元数据
            if options.include_metadata:
                metadata = [
                    ("导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                    ("文档类型", str(document.type)),
                    ("字符数", len(document.content)),
                    ("词数", len(document.content.split())),
                ]
                
                for i, (label, value) in enumerate(metadata, 2):
                    ws[f'A{i}'] = label
                    ws[f'B{i}'] = value
                    ws[f'A{i}'].font = Font(bold=True)
                    
                # 添加分隔行
                ws[f'A{len(metadata) + 3}'] = "文档内容"
                ws[f'A{len(metadata) + 3}'].font = Font(bold=True)
                content_start_row = len(metadata) + 4
            else:
                content_start_row = 2
            
            # 添加文档内容（按段落分行）
            paragraphs = document.content.split('\n\n')
            for i, paragraph in enumerate(paragraphs):
                if paragraph.strip():
                    ws[f'A{content_start_row + i}'] = paragraph.strip()
            
            # 调整列宽
            ws.column_dimensions['A'].width = 80
            ws.column_dimensions['B'].width = 30
            
            # 保存工作簿
            wb.save(str(output_path))
            
            logger.info(f"文档已导出为Excel: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出文档为Excel失败: {e}")
            return False

    async def _do_import_document(self, input_path: Path, options: ImportOptions) -> Optional[Document]:
        """从Excel导入文档"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl库不可用，无法导入Excel格式")
            return None
            
        try:
            # 验证文件扩展名
            if not self._validate_file_extension(input_path, self.get_supported_extensions()):
                logger.error(f"不支持的文件格式: {input_path.suffix}")
                return None

            # 读取Excel文件
            wb = openpyxl.load_workbook(str(input_path))
            ws = wb.active
            
            # 提取标题（第一行）
            title = input_path.stem
            if ws['A1'].value:
                title = str(ws['A1'].value)
            
            # 提取内容
            content_parts = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and str(row[0]).strip():
                    content_parts.append(str(row[0]).strip())
            
            content = '\n\n'.join(content_parts)
            
            # 创建文档
            document = Document(
                id="",  # 将由仓储分配
                title=title,
                content=content.strip(),
                type=DocumentType.CHAPTER
            )
            
            logger.info(f"文档已从Excel导入: {document.title}")
            return document
            
        except Exception as e:
            logger.error(f"从Excel导入文档失败: {e}")
            return None
